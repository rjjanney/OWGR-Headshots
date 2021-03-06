from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import requests
from openpyxl.worksheet.table import Table, TableStyleInfo

# Constants
RED = "FFFF8080"
GREEN = "FF80FF80"
EXCEL_FILE = "../OWGR.xlsx"
URL = "http://www.owgr.com/ranking"


def get_excel_sheet_data(excel_file):

    '''
    Reads an excel file and makes a list of pairs of player names
    and OWG ranking.
    '''

    workbook = load_workbook(excel_file, True)
    worksheet = workbook.active
    # print(worksheet.dimensions)
    old_list = []
    for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=5):
        if row[0].value:
            # rank, name, headshot or not
            old_list.append((row[0].value,
                             row[1].value,
                             row[2].fill.bgColor.rgb,
                             row[3].fill.bgColor.rgb))
    return old_list


def get_rank_names_list(url):

    '''
    A function to scrape the top 100 PGA player rankings and names
    from www.owgr.com/ranking, or from a saved html file during testing
    returns a list of player names, last name first, comma separated,
    and current OWG ranking.

    *** for some reason, probably user agent identification, the html saved
    to disk is different than the html scraped live. Specifically, there
    exists a <tbody> tag in the saved file, which isn't there live. Then
    that difference makes an extra 1st entry show up in the live feed, hence
    line 58 (pop the first entry off of the list). ***
    '''

    html = requests.get(url)
    # html = open('docs/Official World Golf Ranking - Ranking.html')
    bsObj = BeautifulSoup(html.content, "html.parser")
    rank_name_list = []
    for item in bsObj.body.find(
                                'div',
                                class_='table_container'
                                ).table.findAll('tr'):

        # print(item.encode('utf8')) # gets rid of ascii encoding error
        rank = item.findNext('td').contents[1]
        # find gender in the next item in the <ul> ... </ul>
        name = item.findNext('td', {'class': 'name'}).a.string

        first_name = name.split(' ')[0]
        last_name_list = name.split(' ')[1:]
        last_name = ' '.join(last_name_list)
        rank_name_list.append((int(rank), last_name + ", " + first_name))
    rank_name_list.pop(0)
    return rank_name_list


def process_lists(old_list, new_list):

    updated_list = []
    # make copy of old_list - working_list
    working_list = list(old_list)
    unmatched_new_list = list(new_list)
    color1 = ""
    color2 = ""

    # compare each line of latest list to each line of working list
    for item in new_list:
        for line in old_list:
            # if names match,
            if item[1].lower() == line[1].lower():
                if line[2] == GREEN:
                    # copy color column1
                    color1 = GREEN

                elif line[2] == RED:

                    color1 = RED

                if line[3] == GREEN:
                    # copy color column1
                    color2 = GREEN

                elif line[3] == RED:

                    color2 = RED

                # remove match from working list and unmatched_new_list
                working_list.remove(line)
                unmatched_new_list.remove(item)
            # copy existing list entry and add rank as 3rd & 4th entry
            # in new list "updated_list"

                updated_list.append((item[0],
                                     item[1],
                                     color1,
                                     color2))

    # append non-matches to updated list with RED formatting
    for item in unmatched_new_list:
        updated_list.append((item[0],
                             item[1],
                             RED,
                             RED))

    # copy all the non-ranked headshots matches
    # to updated_list, with ranking of 101
    # preserving "has head shot" info in line[2 and 3]
    for line in working_list:
        updated_list.append((101,
                             line[1],
                             line[2],
                             line[3]))

    return updated_list


def convert_colors(list_with_colors):
    '''
    Converts color values to excel sheet style fill objects
    and returns new list
    '''

    new_formatted_list = []
    for line in list_with_colors:
        color2017 = PatternFill(start_color=line[2],
                                end_color=line[2],
                                fill_type="solid")
        color2018 = PatternFill(start_color=line[3],
                                end_color=line[3],
                                fill_type="solid")
        new_formatted_list.append((line[0], line[1], color2017, color2018))
    return new_formatted_list


def save_updated_excel_file(updated_list):

    '''
    Takes original excel sheet and formats it for use with the scraped data,
    including adding color to 2nd and 3rd column to show GREEN "80FF80"
    if headshot exists.
    '''
    thin_border = Border(bottom=Side(style='thin'), left=Side(style='thin'))
    thick_border = Border(bottom=Side(style='thick'))

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.column_dimensions["B"].width = 20
    worksheet.cell(3, 2).value = "RANK"
    worksheet.cell(3, 3).value = "Player Name"
    worksheet.cell(3, 4).value = "2017"
    worksheet.cell(3, 5).value = "2018"

    row = 4
    for line in updated_list:
        worksheet.cell(row, 2).value = line[0]
        worksheet.cell(row, 2).font = Font(bold=True, size=14.0)
        worksheet.cell(row, 2).alignment = Alignment(horizontal="center",
                                                     shrinkToFit=True)
        worksheet.cell(row, 3).value = line[1]
        worksheet.cell(row, 3).alignment = Alignment(horizontal="left")
        worksheet.cell(row, 4).border = thin_border
        worksheet.cell(row, 4).fill = line[2]
        worksheet.cell(row, 5).border = thin_border
        worksheet.cell(row, 5).fill = line[3]
        row += 1

    # column widths
    worksheet.column_dimensions["B"].width = 6
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 10
    worksheet.column_dimensions["E"].width = 10

    # thick line for the cutoff rank
    for i in range(6):
        worksheet.cell(67, i+1).border = thick_border

    tab = Table(displayName="Table1", ref=("B3:E" + str(row-1)))
    style = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=False,
                           showColumnStripes=False)
    tab.tableStyleInfo = style
    worksheet.add_table(tab)

    workbook.save('OWGR.xlsx')


def spit_out_headshots_needed(updated_list):
    '''
    Takes an ordered list of (rank, player name, [COLOR]) and
    saves to a text file a title and "rank    player name" for all that
    have the color RED.
    '''

    heads_needed = []
    for line in updated_list:
            # if names match,
            if line[2] == RED and line[3] == RED:
                # copy existing list entry and add rank as a 3rd entry
                # in new list "updated_list"
                heads_needed.append((line[0],
                                     line[1]))
    # pprint.pprint(heads_needed)

    with open('heads_needed.txt', 'w') as file_handler:
        file_handler.write("HEADSHOTS THAT ARE NEEDED\n\n")
        for line in heads_needed:
            file_handler.write("{}  {}\n".format(line[0], line[1]))

    return

if __name__ == "__main__":

    existing_list = get_excel_sheet_data(EXCEL_FILE)
    latest_list = get_rank_names_list(URL)
    up_to_date_list = process_lists(existing_list, latest_list)

    # sort list by rank
    up_to_date_list.sort(key=lambda x: x[0])

    # save text list of heads needed
    spit_out_headshots_needed(up_to_date_list)

    # convert colors to excel cell format
    list_to_save = convert_colors(up_to_date_list)

    # save updated_list to excel file
    save_updated_excel_file(list_to_save)
