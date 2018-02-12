from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pprint

red = "FF8080"
green = "80FF80"


class Player:

    def __init__(self, name, rank):
        self.name = name
        self.rank = rank


def get_rank_names_list():

    '''
    A function to scrape the top 100 PGA player rankings and names
    from www.owgr.com/ranking, or from a saved html file during testing
    returns a list of player names, last name first, comma separated,
    and current OWG ranking.
    '''

    # html = urlopen("http://www.owgr.com/ranking")
    html = open('docs/Official World Golf Ranking - Ranking.html')
    bsObj = BeautifulSoup(html, "html.parser")
    rank_name_list = []
    for item in bsObj.body.find('div', class_='table_container').table.tbody.findAll('tr'):
        # print(item.encode('utf8')) # gets rid of ascii encoding error
        # find age group string inside a <ul><li><span> ... </span></li></ul>
        rank = item.findNext('td').contents[1]
        # find gender in the next item in the <ul> ... </ul>
        name = item.findNext('td', {'class': 'name'}).a.string

        first_name = name.split(' ')[0]
        last_name_list = name.split(' ')[1:]
        last_name = ' '.join(last_name_list)
        rank_name_list.append((last_name + ", " + first_name, int(rank)))

    return rank_name_list


def get_excel_sheet_data():

    '''
    Reads an excel file and makes a list of pairs of player names
    and OWG ranking.
    '''

    workbook = load_workbook('docs/test_fixed.xlsx')
    worksheet = workbook.active
    # print(worksheet.dimensions)
    old_list = []
    for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=3):
        if row[0].value:
            old_list.append((row[0].value, row[1].fill.bgColor.rgb))

    return old_list


def parse_names(messed_up_names_list):

    '''
    Function originally used to reformat the excel sheet list
    so that the name formatting matched the data scraped from
    OWGR.com

    Takes list of pairs of names and fill.bgColor.rgb value, used
    to show if headshot exists or not
    '''

    fixed_list = []
    for name, existing in messed_up_names_list:
        if name:
            last_name = name.split(',')[0].capitalize()
            first_name = name.split(',')[1]
            fixed_list.append((last_name + "," + first_name, existing))

    return fixed_list


def save_starting_list(adjusted_list):

    '''
    Takes original excel sheet and formats it for use with the scraped data,
    including adding color to 2nd column to show green "80FF80" if headshot
    exists.
    '''
    workbook = load_workbook('docs/test_fixed.xlsx')
    worksheet = workbook.active

    row = 4
    for line in adjusted_list:
        worksheet.cell(row, 2).value = line[0]
        worksheet.cell(row, 3).fill = PatternFill(start_color=green,
                                                  end_color=green,
                                                  fill_type="solid")
        row += 1

    workbook.save('docs/new_starting_point.xlsx')


def save_updated_excel_file(updated_list):

    '''
    Takes original excel sheet and formats it for use with the scraped data,
    including adding color to 2nd column to show green "80FF80" if headshot
    exists.
    '''
    workbook = Workbook()
    worksheet = workbook.active

    row = 4
    for line in updated_list:
        worksheet.cell(row, 2).value = line[0]
        worksheet.cell(row, 3).fill = line[1]
        worksheet.cell(row, 4).value = line[2]
        row += 1

    workbook.save('docs/test_done.xlsx')


latest_list = get_rank_names_list()
existing_list = get_excel_sheet_data()

# adjusted_list = parse_names(existing_list)
# save_starting_list(adjusted_list)

# pprint.pprint(latest_list)
# pprint.pprint(existing_list)

updated_list = []
match_list = []
# make copy of existing_list - working_list
working_list = list(existing_list)
unmatched_latest_list = list(latest_list)

# compare each line of latest list to each line of working list
for item in latest_list:
    # pprint.pprint(item)
    for line in existing_list:
        # pprint.pprint(line)
        # if names match,
        if item[0].lower() == line[0].lower():
            # copy existing list entry and add rank as a 3rd entry
            # in new list "updated_list"
            updated_list.append((item[0],
                                 PatternFill(start_color=green,
                                             end_color=green,
                                             fill_type="solid"),
                                 item[1]))
            # add match to match_list
            working_list.remove(line)
            unmatched_latest_list.remove(item)

# append non-matches to updated list with red formatting
for item in unmatched_latest_list:
    updated_list.append((item[0],
                        PatternFill(start_color=red,
                                    end_color=red,
                                    fill_type="solid"),
                        item[1]))

# copy all the non-ranked headshots matches
# to updated_list, with ranking of 101
for line in working_list:
    updated_list.append((line[0],
                         PatternFill(start_color=green,
                                     end_color=green,
                                     fill_type="solid"),
                         101))

# sort list by rank

updated_list.sort(key=lambda x: x[2])

# save updated_list to excel file

# pprint.pprint(updated_list)
save_updated_excel_file(updated_list)
