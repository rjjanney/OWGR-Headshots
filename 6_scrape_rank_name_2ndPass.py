from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import requests

red = "00FF8080"
green = "0080FF80"


def get_rank_names_list():

    '''
    A function to scrape the top 100 PGA player rankings and names
    from www.owgr.com/ranking, or from a saved html file during testing
    returns a list of player names, last name first, comma separated,
    and current OWG ranking.
    '''

    url = "http://www.owgr.com/ranking"
    html = requests.get(url)
    # html = open('docs/Official World Golf Ranking - Ranking.html')
    bsObj = BeautifulSoup(html.content, "html.parser")
    rank_name_list = []
    for item in bsObj.body.find('div', class_='table_container').table.findAll('tr'):
        # print(item.encode('utf8')) # gets rid of ascii encoding error
        rank = item.findNext('td').contents[1]
        # find gender in the next item in the <ul> ... </ul>
        name = item.findNext('td', {'class': 'name'}).a.string

        first_name = name.split(' ')[0]
        last_name_list = name.split(' ')[1:]
        last_name = ' '.join(last_name_list)
        rank_name_list.append((last_name + ", " + first_name, int(rank)))
    rank_name_list.pop(0)
    return rank_name_list


def get_excel_sheet_data():

    '''
    Reads an excel file and makes a list of pairs of player names
    and OWG ranking.
    '''

    workbook = load_workbook('docs/test_done.xlsx')
    worksheet = workbook.active
    # print(worksheet.dimensions)
    old_list = []
    for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=4):
        if row[0].value:
            old_list.append((row[0].value, row[1].fill.bgColor.rgb, row[2].value))
    return old_list


def save_updated_excel_file(updated_list):

    '''
    Takes original excel sheet and formats it for use with the scraped data,
    including adding color to 2nd column to show green "80FF80" if headshot
    exists.
    '''
    workbook = Workbook()
    worksheet = workbook.active

    row = 4
    for line in updated_list:
        worksheet.cell(row, 2).value = line[0]
        worksheet.cell(row, 3).fill = line[1]
        worksheet.cell(row, 4).value = line[2]
        row += 1

    workbook.save('docs/test_2ndWeek.xlsx')


latest_list = get_rank_names_list()
existing_list = get_excel_sheet_data()


updated_list = []
match_list = []
# make copy of existing_list - working_list
working_list = list(existing_list)
unmatched_latest_list = list(latest_list)

# compare each line of latest list to each line of working list
for item in latest_list:
    # pprint.pprint(item)
    for line in existing_list:
        # pprint.pprint(line)
        # if names match,
        if item[0].lower() == line[0].lower():
            if line[1] == green:
                # copy existing list entry and add rank as a 3rd entry
                # in new list "updated_list"
                updated_list.append((item[0],
                                     PatternFill(start_color=green,
                                                 end_color=green,
                                                 fill_type="solid"),
                                     item[1]))
            elif line[1] == red:
                # copy existing list entry and add rank as a 3rd entry
                # in new list "updated_list"
                updated_list.append((item[0],
                                     PatternFill(start_color=red,
                                                 end_color=red,
                                                 fill_type="solid"),
                                     item[1]))
            # remove match from working list and unmatched_latest_list
            working_list.remove(line)
            unmatched_latest_list.remove(item)

# append non-matches to updated list with red formatting
for item in unmatched_latest_list:
    updated_list.append((item[0],
                        PatternFill(start_color=red,
                                    end_color=red,
                                    fill_type="solid"),
                        item[1]))

# copy all the non-ranked headshots matches
# to updated_list, with ranking of 101
for line in working_list:
    updated_list.append((line[0],
                         PatternFill(start_color=line[1],
                                     end_color=line[1],
                                     fill_type="solid"),
                         101))

# sort list by rank

updated_list.sort(key=lambda x: x[2])

# save updated_list to excel file
# pprint.pprint(updated_list)
save_updated_excel_file(updated_list)
