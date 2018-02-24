from OWGR import OWGR
from openpyxl import load_workbook


def test_get_excel_sheet_data():
    # test that I can read the desired data from an excel file and return
    # a list
    results = OWGR.get_excel_sheet_data("docs/test.xlsx")
    # assert that they equal my expectations
    assert results == [(1, "Johnson, Dustin", "FF80FF80", "FFFF8080"),
                       (2, "Dostoyevsky, Fyodor", "FF80FF80", "FFFF8080")]


def test_process_lists():
    # test that I can compare two lists and return a third list
    from_excel = [
                  (1, "Johnson, Dustin", "FF80FF80", "FFFF8080"),
                  (2, "Rahm, Jon", "FF80FF80", "FFFF8080")
                  ]
    from_web = [
                (1, "Johnson, Dustin"),
                (2, "Kingsley, Hank"),
                (3, "Rahm, Jon"),
                (4, "Feeble, Hinkey")
                ]
    from_function = [(1, 'Johnson, Dustin', "FF80FF80", "FFFF8080"),
                     (3, 'Rahm, Jon', "FF80FF80", "FFFF8080"),
                     (2, 'Kingsley, Hank', "FFFF8080", "FFFF8080"),
                     (4, 'Feeble, Hinkey', "FFFF8080", "FFFF8080")]
    results = OWGR.process_lists(from_excel, from_web)

    assert results == from_function


def test_convert_colors():
    ''' test if colors get converted into excel things
    '''
    workbook = load_workbook("docs/test.xlsx", True)
    worksheet = workbook.active
    # print(worksheet.dimensions)
    for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=5):
        value_from_file = row[2]
        break

    print(value_from_file.fill)

    list_to_function = [(1, 'Johnson, Dustin', "FF80FF80", "FFFF8080"),
                        (3, 'Rahm, Jon', "FF80FF80", "FFFF8080"),
                        (2, 'Kingsley, Hank', "FFFF8080", "FFFF8080"),
                        (4, 'Feeble, Hinkey', "FFFF8080", "FFFF8080")]
    from_function = OWGR.convert_colors(list_to_function)

    assert(from_function[0][2] == value_from_file.fill)


def test_save_updated_excel_file():
    ''' test that the colors are getting encoded right and the
        excel file is saving correctly
        '''
    assert 0 == 1
