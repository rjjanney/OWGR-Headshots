"""
OWGR Headshot Have / Need list - Writes to Excel file.

Python 3
"""

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import requests
from openpyxl.worksheet.table import Table, TableStyleInfo
# import pprint


red = "FFFF8080"
green = "FF80FF80"


def get_excel_sheet_data():
    """Read excel file, make list of pairs of player names and OWGRs."""
    workbook = load_workbook('OWGR.xlsx', True)
    worksheet = workbook.active
    # print(worksheet.dimensions)
    old_list = []
    for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=7):
        if row[0].value:
            # rank, name, headshot or not
            old_list.append((row[0].value,
                             row[1].value,
                             row[2].fill.bgColor.rgb,
                             row[3].fill.bgColor.rgb,
                             row[4].fill.bgColor.rgb,
                             row[5].fill.bgColor.rgb))
    return old_list


def get_rank_names_list():
    """
    A function to scrape the top 100 PGA player rankings and names.

    From www.owgr.com/ranking, or from a saved html file during testing
    returns a list of player names, last name first, comma separated,
    and current OWG ranking.

    *** for some reason, probably user agent identification, the html saved
    to disk is different than the html scraped live. Specifically, there
    exists a <tbody> tag in the saved file, which isn't there live. Then
    that difference makes an extra 1st entry show up in the live feed, hence
    line 58 (pop the first entry off of the list). ***
    """
    url = "http://www.owgr.com/ranking"
    html = requests.get(url)
    # html = open('docs/Official World Golf Ranking - Ranking.html')
    bsObj = BeautifulSoup(html.content, "html.parser")
    rank_name_list = []
    for item in bsObj.body.find(
            'div',
            class_='table_container').table.findAll('tr'):

        # print(item.encode('utf8')) # gets rid of ascii encoding error
        rank = item.findNext('td').contents[1]
        # find gender in the next item in the <ul> ... </ul>
        name = item.findNext('td', {'class': 'name'}).a.string

        first_name = name.split(' ')[0]
        last_name_list = name.split(' ')[1:]
        last_name = ' '.join(last_name_list)
        rank_name_list.append((int(rank), last_name + ", " + first_name))
    rank_name_list.pop(0)
    return rank_name_list


def process_lists(old_list, new_list):

    updated_list = []
    # make copy of old_list - working_list
    working_list = list(old_list)
    unmatched_new_list = list(new_list)
    # Color 1, 2, 3 == 2017, 2018, 2019, etc.
    color1 = ""
    color2 = ""
    color3 = ""
    color4 = ""

    # compare each line of latest list to each line of working list
    for item in new_list:
        for line in old_list:

            # if names match,
            if item[1].lower() == line[1].lower():
                if line[2] == green:
                    # copy color column1
                    color1 = green

                elif line[2] == red:

                    color1 = red

                if line[3] == green:
                    # copy color column1
                    color2 = green

                elif line[3] == red:

                    color2 = red

                if line[4] == green:
                    # copy color column1
                    color3 = green

                elif line[4] == red:

                    color3 = red

                if line[5] == green:
                    # copy color column1
                    color4 = green

                elif line[5] == red:

                    color4 = red

                # remove match from working list and unmatched_new_list
                working_list.remove(line)
                if item in unmatched_new_list:
                    unmatched_new_list.remove(item)
            # copy existing list entry and add rank as 3rd & 4th entry
            # in new list "updated_list"
        updated_list.append((item[0],
                             item[1],
                             PatternFill(start_color=color1,
                                         end_color=color1,
                                         fill_type="solid"),
                             PatternFill(start_color=color2,
                                         end_color=color2,
                                         fill_type="solid"),
                             PatternFill(start_color=color3,
                                         end_color=color3,
                                         fill_type="solid"),
                             PatternFill(start_color=color4,
                                         end_color=color4,
                                         fill_type="solid")))

    # append non-matches to updated list with red formatting
    for item in unmatched_new_list:
        updated_list.append((item[0],
                             item[1],
                             PatternFill(start_color=red,
                                         end_color=red,
                                         fill_type="solid"),
                             PatternFill(start_color=red,
                                         end_color=red,
                                         fill_type="solid"),
                             PatternFill(start_color=red,
                                         end_color=red,
                                         fill_type="solid"),
                             PatternFill(start_color=red,
                                         end_color=red,
                                         fill_type="solid")))

    # copy all the non-ranked headshots matches
    # to updated_list, with ranking of 101
    # preserving "has head shot" info in line[1]
    for line in working_list:
        updated_list.append((101,
                             line[1],
                             PatternFill(start_color=line[2],
                                         end_color=line[2],
                                         fill_type="solid"),
                             PatternFill(start_color=line[3],
                                         end_color=line[3],
                                         fill_type="solid"),
                             PatternFill(start_color=line[4],
                                         end_color=line[4],
                                         fill_type="solid"),
                             PatternFill(start_color=line[5],
                                         end_color=line[5],
                                         fill_type="solid")))
    return updated_list


def save_updated_excel_file(updated_list):
    """
    Take original excel sheet and format it.

    For use with the scraped data,
    including adding color to 2nd column to show green "80FF80" if headshot
    exists.
    """
    thin_border = Border(bottom=Side(style='thin'), left=Side(style='thin'))
    thick_border = Border(bottom=Side(style='thick'))

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.column_dimensions["B"].width = 20
    worksheet.cell(3, 2).value = "RANK"
    worksheet.cell(3, 3).value = "Player Name"
    worksheet.cell(3, 4).value = "2017"
    worksheet.cell(3, 5).value = "2018"
    worksheet.cell(3, 6).value = "2019"
    worksheet.cell(3, 7).value = "2020"

    row = 4
    for line in updated_list:
        worksheet.cell(row, 2).value = line[0]
        worksheet.cell(row, 2).font = Font(bold=True, size=14.0)
        worksheet.cell(row, 2).alignment = Alignment(horizontal="center",
                                                     shrinkToFit=True)
        worksheet.cell(row, 3).value = line[1]
        worksheet.cell(row, 3).alignment = Alignment(horizontal="left")
        worksheet.cell(row, 4).border = thin_border
        worksheet.cell(row, 4).fill = line[2]
        worksheet.cell(row, 5).border = thin_border
        worksheet.cell(row, 5).fill = line[3]
        worksheet.cell(row, 6).border = thin_border
        worksheet.cell(row, 6).fill = line[4]
        worksheet.cell(row, 7).border = thin_border
        worksheet.cell(row, 7).fill = line[5]
        row += 1

    # column widths
    worksheet.column_dimensions["B"].width = 6
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 10
    worksheet.column_dimensions["E"].width = 10
    worksheet.column_dimensions["F"].width = 10
    worksheet.column_dimensions["G"].width = 10

    # thick line for the cutoff rank
    for i in range(8):
        worksheet.cell(67, i+1).border = thick_border

    tab = Table(displayName="Table1", ref=("B3:F" + str(row-1)))
    style = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=False,
                           showColumnStripes=False)
    tab.tableStyleInfo = style
    worksheet.add_table(tab)

    workbook.save('OWGR.xlsx')


def spit_out_headshots_needed(updated_list):
    """
    Separate Needed Heads.

    Takes an ordered list of (rank, player name, [COLOR]) and
    saves to a text file a title and "rank player name" for all that
    have the color red.
    """
    heads_needed = []
    for line in updated_list:
            # If red, indicating missing headshot, in 2019 column(line[4])
            if line[4].bgColor.rgb == red:
                # copy existing rank and name as an entry
                # in new list "heads needed"
                heads_needed.append((line[0],
                                     line[1]))
    # pprint.pprint(heads_needed)

    with open('heads_needed.txt', 'w') as file_handler:
        file_handler.write("HEADSHOTS THAT WE NEED\n\n")
        for line in heads_needed:
            file_handler.write("{}  {}\n".format(line[0], line[1]))

    return

existing_list = get_excel_sheet_data()
latest_list = get_rank_names_list()
up_to_date_list = process_lists(existing_list, latest_list)

# sort list by rank
up_to_date_list.sort(key=lambda x: x[0])

# save text list of heads needed
spit_out_headshots_needed(up_to_date_list)

# save updated_list to excel file
save_updated_excel_file(up_to_date_list)
